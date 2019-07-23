import requests
from bs4 import BeautifulSoup, Comment
from openpyxl import load_workbook, Workbook
import os
import time
import traceback
from datetime import datetime
from configparser import ConfigParser
from abc import ABCMeta, abstractmethod
from urllib.parse import urlparse, urljoin

PAGE = 10
TIMEOUT = 1


def get_cur_time_filename():
    return time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())


def format_cd_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d小时%02d分%02d秒" % (h, m, s)


def get_url_domain(url):
    li = urlparse(url).netloc.split('.')
    length = len(li)
    return '{}.{}'.format(li[length - 2], li[length - 1])


class MyError(RuntimeError):
    pass


class SpiderRuler(metaclass=ABCMeta):
    def __init__(self, spider):
        self.spider = spider

    @property
    @abstractmethod
    def user_agent(self):
        pass

    @property
    @abstractmethod
    def base_url(self):
        pass

    @property
    @abstractmethod
    def engine_name(self):
        pass

    @property
    @abstractmethod
    def request_interval_time(self):
        pass

    @abstractmethod
    def get_params(self, keyword, page):
        pass

    def check_url(self, url):
        return True, ''

    @abstractmethod
    def get_all_item(self, soup):
        pass

    @abstractmethod
    def get_url(self, item):
        pass

    @abstractmethod
    def get_title(self, item):
        pass

    @abstractmethod
    def has_next_page(self, soup):
        pass


class SMRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'sm_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'

    @property
    def base_url(self):
        return 'https://m.sm.cn/s'

    @property
    def engine_name(self):
        return '神马'

    def get_params(self, keyword, page):
        return {
            'q': keyword,
            'page': page,
            'by': 'next',
            'from': 'smor',
            'tomode': 'center',
            'safe': '1',
        }

    def get_all_item(self, soup):
        return soup.find_all('div', class_='ali_row')

    def get_url(self, item):
        link = item.find('a')
        return link.get('href')

    def get_title(self, item):
        return ''.join(item.find('a').findAll(text=True))

    def has_next_page(self, soup):
        return soup.find('a', class_='next')


class SogouPCRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'sgpc_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36'

    @property
    def base_url(self):
        return 'http://www.sogou.com/web'

    @property
    def engine_name(self):
        return '搜狗PC'

    def get_params(self, keyword, page):
        return {
            'query': keyword,
            'page': page,
        }

    def check_url(self, url):
        if url.startswith('http://www.sogou.com/antispider'):
            return False, '该IP已经被搜狗引擎封禁'
        return True, ''

    def get_all_item(self, soup):
        return soup.find('div', class_='results').find_all('div', recursive=False)

    def get_url(self, item):
        link = item.find('a')
        return link.get('href')

    def get_title(self, item):
        return ''.join(item.find('a').findAll(text=lambda text: not isinstance(text, Comment)))

    def has_next_page(self, soup):
        return soup.find(id='sogou_next')


class SogouMobileRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'sgmobile_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Mobile Safari/537.36'

    @property
    def base_url(self):
        return 'http://wap.sogou.com/web/search/ajax_query.jsp'

    @property
    def engine_name(self):
        return '搜狗MOBILE'

    def get_params(self, keyword, page):
        return {
            'keyword': keyword,
            'p': page,
        }

    def get_all_item(self, soup):
        return soup.find_all('a', class_='resultLink')

    def get_url(self, item):
        url = item.get('href')
        if url.startswith('javascript'):
            return None
        elif url.startswith('http'):
            return url
        else:
            url = urljoin(self.spider.url, url)
            (r, sub_soup) = self.spider.safe_request(url)
            if r.url.startswith('http://wap.sogou.com/web/search'):
                btn = sub_soup.find('div', class_='btn')
                if btn:
                    link = btn.find('a')
                else:
                    # 个别情况下 会发生页面里面没有class为btn的div的情况
                    link = sub_soup.find('a')
                return link.get('href')
            else:
                return r.url

    def get_title(self, item):
        return ''.join(item.findAll(text=lambda text: not isinstance(text, Comment)))

    def has_next_page(self, soup):
        li = soup.find('p').find(text=True).strip().split(',')
        return int(li[0]) > int(li[1])


class Spider(metaclass=ABCMeta):
    def __init__(self, ruler_class):
        self.ruler = ruler_class(self)
        self.url = ''
        self.text = ''
        self.result = []
        self.started = False
        self.last_request_time = datetime.now()

    def main(self):
        try:
            self.get_mode()
        except MyError as e:
            self.save_result()
            input(e)
        except KeyboardInterrupt:
            self.save_result()
            input('已经强行退出程序')
        except:
            self.save_result()
            filename = 'error-%s.log' % get_cur_time_filename()
            with open(filename, 'w', encoding='utf-8') as f:
                f.write('''%s

请求的URL为：
%s

返回的内容为：
%s
''' % (traceback.format_exc(), self.url, self.text))
            traceback.print_exc()
            input('请将最新的error.log文件发给技术人员')

    def get_mode(self):
        while True:
            run_mode = input('定时运行（输入1）还是马上运行（输入0）？')
            if run_mode == '0':
                self.search()
            elif run_mode == '1':
                self.start()
            else:
                print('输入了未知模式，请重新输入')

    def start(self):
        now = datetime.now()
        cfg = ConfigParser()
        cfg.read('config.ini')
        hour = int(cfg.get('config', 'hour'))
        start_time = datetime(now.year, now.month, now.day, hour)
        if start_time <= now:
            start_time = datetime.fromtimestamp(start_time.timestamp() + 24 * 60 * 60)
        wait_time = (start_time - now).total_seconds()
        print('下次查询时间为%s，将在%s后开始' % (start_time, format_cd_time(wait_time)))
        time.sleep(wait_time)
        self.search()
        self.start()

    @abstractmethod
    def search(self):
        pass

    @abstractmethod
    def save_result(self):
        pass

    def safe_request(self, url, *, params=None, headers=None):
        cur = datetime.now()
        passed = (cur - self.last_request_time).total_seconds()
        if passed < self.ruler.request_interval_time:
            time.sleep(self.ruler.request_interval_time - passed)
        r = None
        soup = None
        while r is None or soup is None:
            try:
                r = requests.get(url, params=params, headers=headers)
            except (requests.exceptions.ConnectionError, requests.exceptions.ChunkedEncodingError):
                print('检查到网络断开，%s秒之后尝试重新抓取' % TIMEOUT)
                time.sleep(TIMEOUT)
                continue
            self.url = r.url
            self.text = r.text
            soup = BeautifulSoup(r.text, 'lxml')
            if soup.body is None:
                raise MyError('请求到的页面的内容为空，为防止IP被封禁，已退出程序')
        return r, soup


class RankSpider(Spider):
    def __init__(self, ruler_class):
        Spider.__init__(self, ruler_class)
        self.keyword_set = set()
        self.domain_set = set()
        self.searched_keywords = []
        self.main()

    def search(self):
        self.started = True
        self.result = []
        self.searched_keywords = []
        self.keyword_set = set()
        self.domain_set = set()

        start_time = datetime.now()
        (keyword_set, domain_set) = self.get_input()
        self.keyword_set = keyword_set
        self.domain_set = domain_set
        print('总共要查找%s关键词，有%s个网站' % (len(keyword_set), len(domain_set)))
        for i, keyword in enumerate(keyword_set):
            self.get_rank(i + 1, keyword, domain_set)
        self.save_result()
        end_time = datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))
        self.started = False

    def get_input(self):
        file_path = ''
        path = '.\\import'
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            break
        if file_path == '' or not file_path.endswith('.xlsx'):
            raise MyError('import目录之下没有发现xlsx文件')
        wb = load_workbook(file_path)
        ws = wb.active
        d = set()
        k = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                d.add(row[0])
            if row[1] is not None:
                k.add(row[1])
        return k, d

    def get_rank(self, index, keyword, domain_set):
        print('开始抓取第%s个关键词：%s' % (index, keyword))
        for i in range(PAGE):
            self.get_page(i + 1, keyword, domain_set)
        self.searched_keywords.append(keyword)

    def get_page(self, page, keyword, domain_set):
        print('开始第%d页' % page)
        params = self.ruler.get_params(keyword, page)
        headers = {'User-Agent': self.ruler.user_agent}
        (r, soup) = self.safe_request(self.ruler.base_url, params=params, headers=headers)
        (ok, msg) = self.ruler.check_url(r.url)
        if not ok:
            raise MyError(msg)
        all_item = self.ruler.get_all_item(soup)
        rank = 1
        for item in all_item:
            url = self.ruler.get_url(item)
            if url is not None:
                domain = get_url_domain(url)
                if domain in domain_set:
                    self.result.append((
                        domain,
                        keyword,
                        '%d-%d' % (page, rank),
                        url,
                        self.ruler.get_title(item),
                        datetime.now()
                    ))
            rank += 1

    def save_result(self):
        if not self.started:
            return
        file_name = '关键词排名-%s.xlsx' % get_cur_time_filename()
        wb = Workbook()
        ws = wb.active
        ws.append(('域名', '关键词', '搜索引擎', '排名', '真实地址', '标题', '查询时间'))
        for (domain, keyword, rank, url, title, date_time) in self.result:
            ws.append((domain, keyword, self.ruler.engine_name, rank, url, title, date_time))
        wb.save(file_name)
        self.result = []
        print('查询结束，查询结果保存在 %s' % file_name)
        self.save_un_searched()

    def save_un_searched(self):
        un_searched_keywords = []
        for keyword in self.keyword_set:
            if keyword not in self.searched_keywords:
                un_searched_keywords.append(keyword)
        if len(un_searched_keywords) != 0:
            file_name = '未查找关键词-%s.xlsx' % get_cur_time_filename()
            wb = Workbook()
            ws = wb.active
            for keyword in un_searched_keywords:
                ws.append((keyword,))
            wb.save(file_name)
            print('未查询结果保存在 %s' % file_name)


class SiteSpider(Spider):
    def __init__(self, ruler_class):
        Spider.__init__(self, ruler_class)
        self.domain = ''
        self.main()

    def search(self):
        self.started = True
        self.result = []
        start_time = datetime.now()
        self.get_input()
        print('本次查找的域名为 %s' % self.domain)
        page = 1
        soup = self.get_page(self.domain, page)
        while soup and self.ruler.has_next_page(soup):
            page += 1
            soup = self.get_page(self.domain, page)
        self.save_result()
        end_time = datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))
        self.started = False

    def get_input(self):
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.domain = cfg.get('config', 'domain')

    def get_page(self, domain, page):
        print('开始第%d页' % page)
        params = self.ruler.get_params('site:%s' % domain, page)
        headers = {'User-Agent': self.ruler.user_agent}
        (r, soup) = self.safe_request(self.ruler.base_url, params=params, headers=headers)
        (ok, msg) = self.ruler.check_url(r.url)
        if not ok:
            raise MyError(msg)
        all_item = self.ruler.get_all_item(soup)
        for item in all_item:
            self.result.append(self.ruler.get_title(item))
        return soup

    def save_result(self):
        if not self.started:
            return
        wb = Workbook()
        ws = wb.active
        for title in self.result:
            ws.append((title,))
        self.result = []
        file_name = '收录标题-%s.xlsx' % get_cur_time_filename()
        wb.save(file_name)


if __name__ == '__main__':
    spider_list = [(RankSpider, '排名'), (SiteSpider, '收录')]
    engine_list = [(SMRuler, '神马'), (SogouPCRuler, '搜狗PC'), (SogouMobileRuler, '搜狗MOBILE')]
    spider_index = input('''要查找什么数据？
%s
''' % '\n'.join(['%s 请输入：%s' % (ruler_name, i) for (i, (_, ruler_name)) in enumerate(spider_list)]))
    engine_index = input('''要查找哪个搜索引擎？
%s
''' % '\n'.join(['%s 请输入：%s' % (ruler_name, i) for (i, (_, ruler_name)) in enumerate(engine_list)]))
    (spider_class, spider_name) = spider_list[int(spider_index)]
    (ruler_class, ruler_name) = engine_list[int(engine_index)]
    os.system('title %s%s' % (ruler_name, spider_name))
    spider_class(ruler_class)
