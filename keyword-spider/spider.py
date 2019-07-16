import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import time
import datetime
import traceback
import os
from configparser import ConfigParser

TIMEOUT = 1


def get_cur_time_filename():
    return time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())


def format_cd_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d小时%02d分%02d秒" % (h, m, s)


class MyError(RuntimeError):
    pass


class SiteSpider:
    def __init__(self):
        self.file_name = ''
        try:
            while True:
                mode = input('定时运行（输入1）还是马上运行（输入0）？')
                if mode == '0':
                    self.search()
                elif mode == '1':
                    self.start()
                else:
                    print('输入了未知模式，请重新输入')
        except MyError as e:
            input(e)
        except KeyboardInterrupt:
            input('已经强行退出程序')
        except:
            filename = 'error-%s.log' % get_cur_time_filename()
            f = open(filename, 'w', encoding='utf-8')
            f.write(traceback.format_exc())
            f.close()
            traceback.print_exc()
            input('请将最新的error.log文件发给技术人员')

    def search(self):
        start_time = datetime.datetime.now()
        self.file_name = '收录网站标题-%s.xlsx' % get_cur_time_filename()
        wb = Workbook()
        wb.save(self.file_name)
        domain = self.get_input()
        page = 1
        soup = self.get_title(domain, page)
        while soup and soup.find('a', class_='next'):
            page += 1
            soup = self.get_title(domain, page)
        print('查询结束，查询结果保存在 %s' % self.file_name)
        end_time = datetime.datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))

    def start(self):
        now = datetime.datetime.now()
        cfg = ConfigParser()
        cfg.read('config.ini')
        hour = int(cfg.get('config', 'hour'))
        x = datetime.datetime(now.year, now.month, now.day, hour)
        if x <= now:
            x = datetime.datetime.fromtimestamp(x.timestamp() + 24 * 60 * 60)
        wait_time = (x - now).total_seconds()
        print('下次查询时间为%s，将在%s后开始' % (x, format_cd_time(wait_time)))
        time.sleep(wait_time)
        self.search()
        self.start()

    def get_input(self):
        file_path = ''
        path = '.\\import'
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            break
        if file_path == '' or not file_path.endswith('.xlsx'):
            raise MyError('import目录之下没有发现xlsx文件')
        wb = load_workbook(file_path)
        ws = wb.active
        for (domain,) in ws.iter_rows(values_only=True):
            return domain

    def get_title(self, domain, page):
        print('开始抓取第%s页' % page)
        url = 'http://m.sm.cn/s'
        params = {
            'q': 'site:%s' % domain,
            'page': page,
            'by': 'next',
            'from': 'smor',
            'tomode': 'center',
            'safe': '1',
        }
        # todo 不同的UA有什么影响
        # todo 要怎么模拟真实的用户进行搜索 HEAD要怎么填写
        headers = {
            # 'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1'
            'User-Agent': 'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Mobile Safari/537.36'
        }
        r = None
        soup = None
        while r is None:
            try:
                r = requests.get(url, params=params, headers=headers)
            except (requests.exceptions.ConnectionError, requests.exceptions.ChunkedEncodingError):
                print('检查到网络断开，%s秒之后尝试重新抓取' % TIMEOUT)
                time.sleep(TIMEOUT)
                continue
            soup = BeautifulSoup(r.text, 'lxml')
            if soup.body is None:
                print('请求到的页面的内容为空，将再次进行请求')
                r = None
        for child in soup.body.children:
            if child.name == 'div' and child.get('class') and 'ali_row' in child.get('class'):
                title = ''.join(child.find('a').findAll(text=True))
                wb = load_workbook(self.file_name)
                ws = wb.active
                ws.append((title,))
                wb.save(self.file_name)
        return soup


if __name__ == '__main__':
    SiteSpider()
