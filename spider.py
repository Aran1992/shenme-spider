import ast
import os
import re
import sys
import time
import traceback
from abc import ABCMeta, abstractmethod
from configparser import ConfigParser
from datetime import datetime
from urllib.parse import urlparse, parse_qsl, urlsplit, urljoin

import requests
from bs4 import BeautifulSoup, Comment
from openpyxl import load_workbook, Workbook

# import this seems unused
# but it's to prevent 'bs4.FeatureNotFound: Couldn't find a tree builder with the features you requested: lxml.'
import lxml

page_cfg = ConfigParser()
page_cfg.read('config.ini')
PAGE = int(page_cfg.get('config', 'page_count'))


def get_cur_time_filename():
    return time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())


def format_cd_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d小时%02d分%02d秒" % (h, m, s)


def is_list_include_another_list(child_list, parent_list):
    if child_list[0] in parent_list:
        index = parent_list.index(child_list[0])
        for i, child in enumerate(child_list):
            if child not in parent_list or parent_list.index(child) != index + i:
                return False
        return True
    else:
        return False


def page_has_text(soup, text):
    return soup.find(text=re.compile(text))


class MyError(RuntimeError):
    pass


class SpiderRuler(metaclass=ABCMeta):
    def __init__(self, spider):
        self.spider = spider

    @property
    @abstractmethod
    def user_agent(self):
        pass

    @property
    @abstractmethod
    def base_url(self):
        pass

    @property
    @abstractmethod
    def engine_name(self):
        pass

    @property
    @abstractmethod
    def request_interval_time(self):
        pass

    @abstractmethod
    def get_params(self, keyword, page):
        pass

    def is_forbid(self, r, soup):
        return False

    @abstractmethod
    def get_all_item(self, soup):
        pass

    @abstractmethod
    def get_url(self, item, page_url):
        pass

    @abstractmethod
    def get_title(self, item):
        pass

    @abstractmethod
    def has_next_page(self, soup):
        pass

    def get_next_page_url(self, soup):
        return None

    # 默认都开启，有session搜索引擎返回的数据更接近真实情况，而且也比较不容易出错
    @property
    def enable_session(self):
        return True

    @abstractmethod
    def has_no_result(self, soup):
        return False

    def retry_page(self, soup):
        return False

    def is_unsafe(self, item):
        return False


class SMRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'sm_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
               'Chrome/75.0.3770.100 Safari/537.36'

    @property
    def base_url(self):
        return 'https://m.sm.cn/s'

    @property
    def engine_name(self):
        return '神马'

    def get_params(self, keyword, page):
        return {
            'q': keyword,
            'page': page,
            'by': 'next',
            'from': 'smor',
            'tomode': 'center',
            'safe': '1',
        }

    def get_all_item(self, soup):
        return soup.find_all('div', class_='ali_row')

    def get_url(self, item, page_url):
        link = item.find('a')
        return link and link.get('href')

    def get_title(self, item):
        return ''.join(item.find('a').findAll(text=True))

    def has_next_page(self, soup):
        return soup.find('a', class_='next')

    def is_forbid(self, r, soup):
        return soup.body is None \
               or (soup.title and soup.title.text == '验证码拦截')

    def has_no_result(self, soup):
        return page_has_text(soup, '1. 看看输入的文字是否有误') and page_has_text(soup, '2. 去掉可能不必要的字词，如"的"、"什么"等')


class SogouPCRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'sgpc_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
               'Chrome/75.0.3770.100 Safari/537.36'

    @property
    def base_url(self):
        return 'http://www.sogou.com/web'

    @property
    def engine_name(self):
        return '搜狗PC'

    def get_params(self, keyword, page):
        return {
            'query': keyword,
            'page': page,
        }

    def is_forbid(self, r, soup):
        return r.url.startswith('http://www.sogou.com/antispider') \
               or soup.find('div', class_='results') is None

    # results里面可能没有条目，重复请求的话可能又会有了
    def get_all_item(self, soup):
        return soup.find('div', class_='results').find_all('div', recursive=False)

    def get_url(self, item, page_url):
        link = item.find('a')
        return link and link.get('href')

    def get_title(self, item):
        return ''.join(item.find('a').findAll(text=lambda text: not isinstance(text, Comment)))

    def has_next_page(self, soup):
        return soup.find(id='sogou_next')

    def has_no_result(self, soup):
        return soup.find('p', class_='num-tips', text=re.compile('.*?搜狗已为您找到约0条相关结果.*?'))

    # 开启的话 查找到第五页左右搜索引擎就会认为你是爬虫
    @property
    def enable_session(self):
        return False


class SogouMobileRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'sgmobile_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) ' \
               'Chrome/75.0.3770.100 Mobile Safari/537.36'

    @property
    def base_url(self):
        return 'http://wap.sogou.com/web/search/ajax_query.jsp'

    @property
    def engine_name(self):
        return '搜狗MOBILE'

    def get_params(self, keyword, page):
        return {
            'keyword': keyword,
            'p': page,
        }

    def get_all_item(self, soup):
        items = []
        divs = soup.find_all('div', class_='vrResult')
        for div in divs:
            a = div.find('a', class_='resultLink') or div.find('a')
            if a and a.get('href') is not None:
                items.append(a)
        return items

    def get_url(self, item, page_url):
        url = item.get('href')
        if url.startswith('javascript'):
            return None
        elif url.startswith('http'):
            return url
        else:
            url = urljoin(page_url, url)
            query = dict(parse_qsl(urlsplit(url).query))
            if 'url' in query:
                return query['url']
            else:
                (r, sub_soup, _) = self.spider.safe_request(url)
                if r.url.startswith('http://wap.sogou.com/transcoding/sweb') \
                        or r.url.startswith('http://m.sogou.com/transcoding/sweb') \
                        or r.url.startswith('http://wap.sogou.com/web/search/'):
                    btn = sub_soup.find('div', class_='btn')
                    if btn:
                        link = btn.find('a')
                    else:
                        # 个别情况下 会发生页面里面没有class为btn的div的情况
                        link = sub_soup.find('a')
                    if link:
                        return link.get('href')
                else:
                    return r.url

    def get_title(self, item):
        return ''.join(item.findAll(text=lambda text: not isinstance(text, Comment)))

    def has_next_page(self, soup):
        li = soup.find('p').find(text=True).strip().split(',')
        return int(li[0]) > int(li[1])

    def is_forbid(self, r, soup):
        if len(soup.body.contents) == 0:
            return True
        for s in soup.body.contents[0].stripped_strings:
            if s == '403':
                return True

    def has_no_result(self, soup):
        return soup.find('p').find(text=True).strip().split(',')[3] == '0[PAGE_INFO]'


class BaiduPCRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'bdpc_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
               'Chrome/75.0.3770.100 Safari/537.36'

    @property
    def base_url(self):
        return 'https://www.baidu.com/s'

    @property
    def engine_name(self):
        return '百度PC'

    def get_params(self, keyword, page):
        return {
            'wd': keyword,
            'pn': (page - 1) * 10,
        }

    def get_all_item(self, soup):
        div_root = soup.find('div', id='content_left')
        if div_root:
            return div_root.find_all('div', recursive=False, id=lambda id_: id_ != 'rs_top_new')
        else:
            return []

    def get_url(self, item, page_url):
        link = item.find('a')
        if link:
            url = link.get('href')
            if url.startswith('javascript'):
                return None
            elif url.startswith('http://www.baidu.com/link?'):
                return self.spider.get_real_url(url)
            else:
                return url
        else:
            return None

    def get_title(self, item):
        return ''.join(item.find('a').findAll(text=lambda text: not isinstance(text, Comment)))

    def has_next_page(self, soup):
        return soup.find('a', text='下一页 >')

    def get_next_page_url(self, soup):
        page_div = soup.find(id="page")
        if page_div:
            next_page_link = page_div.find('a', text='下一页 >')
            if next_page_link:
                href = next_page_link.get('href')
                if href:
                    return urljoin(self.base_url, href)

    def has_no_result(self, soup):
        return page_has_text(soup, '很抱歉，没有找到与') and page_has_text(soup, '请检查您的输入是否正确')

    def is_unsafe(self, item):
        return item.find('div', class_='unsafe_content f13') is not None

    def is_forbid(self, r, soup):
        # 这种情况其实是等待加载，不算爬虫，但是和爬虫的解决方式是一样的，所以添加在这里
        return r.url.startswith('https://wappass.baidu.com/static/captcha')


class BaiduMobileRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', 'bdmobile_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) ' \
               'Chrome/75.0.3770.100 Mobile Safari/537.36'

    @property
    def base_url(self):
        return 'https://m.baidu.com/s'

    @property
    def engine_name(self):
        return '百度MOBILE'

    def get_params(self, keyword, page):
        return {
            'word': keyword,
            'pn': (page - 1) * 10,
        }

    def get_all_item(self, soup):
        div_root = soup.find('div', id='results')
        if div_root:
            return div_root.find_all('div', class_='c-result result')
        else:
            return []

    def get_url(self, item, page_url):
        data_log_str = item.get('data-log')
        if data_log_str:
            try:
                ast.literal_eval(data_log_str)
                data_log = ast.literal_eval(data_log_str)
                if data_log:
                    mu = data_log.get('mu')
                    if mu and len(mu) > 0:
                        return mu
            except SyntaxError:
                result = re.search(r"'mu':'(.*?)''", data_log_str)
                return result.group(1)

    def get_title(self, item):
        a = item.find('span', class_='c-title-text')
        if a:
            return ''.join(a.findAll(text=lambda text: not isinstance(text, Comment)))
        else:
            return ''

    def has_next_page(self, soup):
        return soup.find('a', class_='new-nextpage-only') or soup.find('a', class_='new-nextpage')

    def get_next_page_url(self, soup):
        next_page_link = soup.find('a', class_='new-nextpage-only') or soup.find('a', class_='new-nextpage')
        if next_page_link:
            href = next_page_link.get('href')
            if href:
                return urljoin(self.base_url, href)

    def is_forbid(self, r, soup):
        # 这种情况其实是等待加载，不算爬虫，但是和爬虫的解决方式是一样的，所以添加在这里
        return r.url.startswith('https://wappass.baidu.com/static/captcha') \
               or (soup.find(id='page-hd') and not soup.find(id='page'))

    def has_no_result(self, soup):
        return page_has_text(soup, '检查输入是否正确') and page_has_text(soup, '抱歉，没有找到与')


class SLLPCRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', '360pc_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) ' \
               'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36'

    @property
    def base_url(self):
        return 'https://www.so.com/s'

    @property
    def engine_name(self):
        return '360PC'

    def get_params(self, keyword, page):
        return {
            'q': keyword,
            'pn': page,
            'src': 'srp_paging',
        }

    def get_all_item(self, soup):
        return soup.find_all(lambda a: a and a.has_attr('data-res'))

    def get_url(self, item, page_url):
        arr = ['data-mdurl', 'data-cache', 'data-url', 'href']
        for key in arr:
            url = item.get(key)
            if url:
                return url

    def get_title(self, item):
        return ''.join(item.findAll(text=lambda text: not isinstance(text, Comment)))

    def has_next_page(self, soup):
        return soup.find('a', id='snext') is not None

    def is_forbid(self, r, soup):
        return page_has_text(soup, '亲，系统检测到您操作过于频繁。')

    def has_no_result(self, soup):
        return page_has_text(soup, '检查输入是否正确') and page_has_text(soup, '简化查询词或尝试其他相关词')

    def retry_page(self, soup):
        return soup.find('ul', class_='result') and len(self.get_all_item(soup)) == 0


class SLLMobileRuler(SpiderRuler):
    def __init__(self, spider):
        SpiderRuler.__init__(self, spider)
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.__request_interval_time = float(cfg.get('config', '360mobile_request_interval_time'))

    @property
    def request_interval_time(self):
        return self.__request_interval_time

    @property
    def user_agent(self):
        return 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) ' \
               'Chrome/78.0.3904.108 Mobile Safari/537.36'

    @property
    def base_url(self):
        return 'https://m.so.com/nextpage'

    @property
    def engine_name(self):
        return '360MOBILE'

    def get_params(self, keyword, page):
        return {
            'q': keyword,
            'src': 'result_input',
            'srcg': 'home_next',
            'pn': page,
            'ajax': 1,
        }

    def get_all_item(self, soup):
        return soup.find_all(lambda div: div and div.has_attr('data-pcurl'))

    def get_url(self, item, page_url):
        return item.get('data-pcurl')

    def get_title(self, item):
        title = item.find('h3', class_='res-title')
        if title:
            return ''.join(title.findAll(text=lambda text: not isinstance(text, Comment)))
        else:
            return ''

    def has_next_page(self, soup):
        return page_has_text(soup, 'MSO.hasNextPage = true;')

    def is_forbid(self, r, soup):
        return page_has_text(soup, '请输入验证码以便正常访问') \
               or r.url.startswith('http://qcaptcha.so.com/?ret=')

    def has_no_result(self, soup):
        return (page_has_text(soup, '很抱歉搜索君没有找到与') and page_has_text(soup, '检查输入是否正确')) \
               or (len(soup.prettify().strip()) == 0) \
               or page_has_text(soup, 'MSO.hasNextPage = false;')


class LittleRankSpider:
    def __init__(self, spider):
        self.spider = spider
        self.error_list = []
        self.page_url = None

    def get_ranks(self, ruler, keyword_domains_map, page):
        result = []
        searched_keywords = []
        for i, keyword in enumerate(keyword_domains_map.keys()):
            domain_set = keyword_domains_map[keyword]
            result += self.get_rank(ruler, i + 1, keyword, domain_set, page)
            searched_keywords.append(keyword)
        return result, self.error_list

    def get_rank(self, ruler, index, keyword, domain_set, page):
        self.spider.reset_session()
        print('开始抓取第%s个关键词：%s' % (index, keyword))
        result = []
        self.page_url = None
        for i in range(page):
            result += self.get_page(ruler, i + 1, keyword, domain_set)
        return result

    def get_page(self, ruler, page, keyword, domain_set):
        print('开始第%d页' % page)
        result = []
        if self.page_url:
            (r, soup, all_item) = self.spider.safe_request(self.page_url)
        else:
            params = ruler.get_params(keyword, page)
            (r, soup, all_item) = self.spider.safe_request(ruler.base_url, params=params)
        self.page_url = ruler.get_next_page_url(soup)
        rank = 1
        for item in all_item:
            try:
                url = ruler.get_url(item, r.url)
            except KeyboardInterrupt as e:
                raise e
            except:
                url = None
                self.error_list.append(traceback.format_exc())
                traceback.print_exc()
            if url is not None:
                print('本页第%s条URL为%s' % (rank, url))
                item_list = urlparse(url).netloc.split('.')
                for domain in domain_set:
                    if domain == '*' or is_list_include_another_list(domain.split('.'), item_list):
                        result.append((
                            domain,
                            keyword,
                            page,
                            rank,
                            url,
                            ruler.get_title(item),
                            datetime.now()
                        ))
                        break
                rank += 1
        return result


class Spider(metaclass=ABCMeta):
    def __init__(self, ruler_class):
        self.ruler = ruler_class(self)
        self.session = None
        self.reset_session()
        self.url = ''
        self.text = ''
        self.result = []
        self.started = False
        self.last_request_time = datetime.now()
        cfg = ConfigParser()
        cfg.read('config.ini')
        self.reconnect_interval_time = float(cfg.get('config', 'reconnect_interval_time'))
        self.error_interval_time = float(cfg.get('config', 'error_interval_time'))
        self.is_keyword_domain_map = int(cfg.get('config', 'is_keyword_domain_map')) == 1
        self.keyword = ''
        self.page = 0

    def main(self):
        try:
            self.get_mode()
        except MyError as e:
            self.save_result()
            input(e)
        except KeyboardInterrupt:
            self.save_result()
            input('已经强行退出程序')
        except:
            self.save_result()
            filename = 'error-%s.log' % get_cur_time_filename()
            with open(filename, 'w', encoding='utf-8') as f:
                f.write('''%s

请求的URL为：
%s

返回的内容为：
%s
''' % (traceback.format_exc(), self.url, self.text))
            traceback.print_exc()
            input('请将最新的error.log文件发给技术人员')

    def get_mode(self):
        first_run = True
        while True:
            if first_run and len(sys.argv) >= 4:
                run_mode = sys.argv[3]
            else:
                run_mode = input('定时运行（输入1）还是马上运行（输入0）？')
            first_run = False
            if run_mode == '0':
                self.search()
            elif run_mode == '1':
                self.start()
            else:
                print('输入了未知模式，请重新输入')

    def start(self):
        now = datetime.now()
        cfg = ConfigParser()
        cfg.read('config.ini')
        hour = int(cfg.get('config', 'hour'))
        start_time = datetime(now.year, now.month, now.day, hour)
        if start_time <= now:
            start_time = datetime.fromtimestamp(start_time.timestamp() + 24 * 60 * 60)
        wait_time = (start_time - now).total_seconds()
        print('下次查询时间为%s，将在%s后开始' % (start_time, format_cd_time(wait_time)))
        time.sleep(wait_time)
        self.search()
        self.start()

    @abstractmethod
    def search(self):
        pass

    @abstractmethod
    def save_result(self):
        pass

    def safe_request(self, url, *, params=None):
        cur = datetime.now()
        passed = (cur - self.last_request_time).total_seconds()
        if passed < self.ruler.request_interval_time:
            time.sleep(self.ruler.request_interval_time - passed)
        r = None
        soup = None
        items = None
        times = 0
        while r is None or soup is None:
            try:
                r = self.get(url, params=params)
            # todo 准确判断是否真的是网络断开 来确定是否要等待网络重连
            except (requests.exceptions.ConnectionError, requests.exceptions.ChunkedEncodingError) as error:
                print('网络断开时请求的URL为：%s' % url)
                print('认为是网络断开的错误是：%s' % error)
                print('检查到网络断开，%s秒之后尝试重新抓取' % self.reconnect_interval_time)
                time.sleep(self.reconnect_interval_time)
                continue
            soup = BeautifulSoup(r.text, 'lxml')
            # with open('1.html', 'w', encoding='utf-8') as f:
            #     f.write(soup.prettify())
            if self.ruler.is_forbid(r, soup):
                # with open(f'旧型爬虫返回页_{self.ruler.engine_name}-{self.keyword}-{self.page}.html',
                #           'w', encoding='utf-8') as f:
                #     f.write(r.url + '\n' + soup.prettify())
                print('该IP已被判定为爬虫，暂时无法获取到信息，%s秒之后尝试重新抓取' % self.error_interval_time)
                time.sleep(self.error_interval_time)
                r = None
                soup = None
                continue
            items = self.ruler.get_all_item(soup)
            if len(items) == 0:
                try:
                    has_no_result = self.ruler.has_no_result(soup)
                except KeyboardInterrupt as e:
                    raise e
                except:
                    has_no_result = False
                if not has_no_result:
                    if not self.ruler.retry_page(soup):
                        with open(f'新型爬虫返回页_可以发送给开发进行分析_{self.ruler.engine_name}-{self.keyword}-{self.page}.html',
                                  'w', encoding='utf-8') as f:
                            f.write(r.url + '\n' + soup.prettify())
                    times = times + 1
                    if times > 5:
                        raise MyError('尝试多次依然无法获取到正确内容')
                    print('请求页面内容异常，可能是被认定为是爬虫，暂时无法获取到信息，%s秒之后尝试重新抓取' % self.error_interval_time)
                    time.sleep(self.error_interval_time)
                    r = None
                    continue
        self.last_request_time = datetime.now()
        self.url = r.url
        self.text = r.text
        return r, soup, items

    def get_real_url(self, start_url):
        cur = datetime.now()
        passed = (cur - self.last_request_time).total_seconds()
        if passed < self.ruler.request_interval_time:
            time.sleep(self.ruler.request_interval_time - passed)
        r = None
        final_url = None
        while r is None:
            try:
                headers = {'User-Agent': self.ruler.user_agent}
                r = requests.head(start_url, headers=headers)
                final_url = r.headers['Location']
            except (requests.exceptions.ConnectionError, requests.exceptions.ChunkedEncodingError):
                print('检查到网络断开，%s秒之后尝试重新抓取' % self.reconnect_interval_time)
                time.sleep(self.reconnect_interval_time)
                continue
            if self.ruler.is_forbid(r, BeautifulSoup(r.text, 'lxml')):
                print('该IP已被判定为爬虫，暂时无法获取到信息，%s秒之后尝试重新抓取' % self.error_interval_time)
                time.sleep(self.error_interval_time)
                continue
        self.last_request_time = datetime.now()
        return final_url

    def reset_session(self):
        if self.ruler.enable_session:
            self.session = requests.Session()
            self.session.headers.update(self.get_headers())

    def get(self, url, *, params=None):
        if self.ruler.enable_session:
            return self.session.get(url, params=params)
        else:
            return requests.get(url, params=params, headers=self.get_headers())

    def get_headers(self):
        return {
            'Accept': 'text/html,application/xhtml+xml,application/xml;'
                      'q=0.9,image/webp,image/apng,*/*;'
                      'q=0.8,application/signed-exchange;'
                      'v=b3',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Pragma': 'no-cache',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': self.ruler.user_agent,
        }


class RankSpider(Spider):
    def __init__(self, ruler_class):
        Spider.__init__(self, ruler_class)
        self.keyword_domains_map = {}
        self.keyword_count = 0
        self.keyword_index = 0
        self.start_time = datetime.now()
        self.searched_keywords = []
        self.filename = ''
        self.error_list = []
        self.unsafe_item_list = []
        self.main()

    def search(self):
        filename_kd_map = self.get_input()
        for index, filename in enumerate(filename_kd_map.keys()):
            self.sub_search(index + 1, filename, filename_kd_map[filename])

    def sub_search(self, index, filename, keyword_domains_map):
        print('开始第%s个文件%s' % (index, filename))
        self.filename = filename
        self.keyword_domains_map = keyword_domains_map
        self.started = True
        self.result = []
        self.searched_keywords = []
        self.start_time = datetime.now()
        self.keyword_count = len(self.keyword_domains_map.keys())
        print('总共要查找%s关键词' % self.keyword_count)
        for i, keyword in enumerate(self.keyword_domains_map.keys()):
            self.keyword_index = i + 1
            self.keyword = keyword
            domain_set = self.keyword_domains_map[keyword]
            self.get_rank(i + 1, keyword, domain_set)
        self.save_result()
        end_time = datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - self.start_time).total_seconds()))
        self.started = False

    def get_input(self):
        filename_kd_map = {}
        path = '.\\import'
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            wb = load_workbook(file_path)
            ws = wb.active
            keyword_domains_map = {}
            if self.is_keyword_domain_map:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    keyword = row[1]
                    domain = row[0]
                    if keyword is not None and domain is not None:
                        if keyword not in keyword_domains_map.keys():
                            keyword_domains_map[keyword] = set()
                        keyword_domains_map[keyword].add(domain)
            else:
                domain_set = set()
                keyword_set = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    keyword = row[1]
                    domain = row[0]
                    if domain is not None:
                        domain_set.add(domain)
                    if keyword is not None:
                        keyword_set.add(keyword)
                for keyword in keyword_set:
                    keyword_domains_map[keyword] = domain_set
            filename_kd_map[file] = keyword_domains_map
        return filename_kd_map

    def get_rank(self, index, keyword, domain_set):
        print('开始抓取第%s个关键词：%s' % (index, keyword))
        self.reset_session()
        page_url = None
        for i in range(PAGE):
            self.page = i + 1
            try:
                page_url, soup = self.get_page(i + 1, keyword, domain_set, page_url)
                if not soup or not self.ruler.has_next_page(soup):
                    break
            except KeyboardInterrupt as e:
                raise e
            except:
                self.error_list.append('关键词：%s，页数：%s，错误：\n%s' % (keyword, i + 1, traceback.format_exc()))
                traceback.print_exc()
        self.searched_keywords.append(keyword)

    def get_page(self, page, keyword, domain_set, page_url):
        print('开始第%d页' % page)
        if page_url:
            (r, soup, all_item) = self.safe_request(page_url)
        else:
            params = self.ruler.get_params(keyword, page)
            (r, soup, all_item) = self.safe_request(self.ruler.base_url, params=params)
        if page == 1:
            try:
                has_no_result = self.ruler.has_no_result(soup)
            except KeyboardInterrupt as e:
                raise e
            except:
                has_no_result = False
            self.unsafe_item_list.append((keyword, (has_no_result and "是") or "否", None, None, None))
        # with open('%s-%s.html' % (keyword, page), 'w', encoding='utf-8') as f:
        #     f.write(soup.prettify())
        print('本页实际请求URL为%s' % r.url)
        os.system('title %s%s 关键词：%s/%s 页数：%s/%s 已用时：%s'
                  % (self.ruler.engine_name, '排名', self.keyword_index, self.keyword_count, page, PAGE,
                     format_cd_time((datetime.now() - self.start_time).total_seconds())))
        rank = 1
        for item in all_item:
            url = self.ruler.get_url(item, r.url)
            if url is not None:
                print('本页第%s条URL为%s' % (rank, url))
                item_list = urlparse(url).netloc.split('.')
                for domain in domain_set:
                    if domain == '*' or is_list_include_another_list(domain.split('.'), item_list):
                        self.result.append((
                            domain,
                            keyword,
                            page,
                            rank,
                            url,
                            self.ruler.get_title(item),
                            datetime.now()
                        ))
                        break
                if self.ruler.is_unsafe(item):
                    self.unsafe_item_list.append((keyword, None, url, page, rank))
                rank += 1
        return self.ruler.get_next_page_url(soup), soup

    def save_result(self):
        if not self.started:
            return
        file_name = '关键词排名-%s-%s-%s.xlsx' % (self.ruler.engine_name, self.filename, get_cur_time_filename())
        wb = Workbook()
        ws = wb.active
        ws.append(('域名', '关键词', '搜索引擎', '页数', '排名', '真实地址', '标题', '查询时间'))
        for (domain, keyword, page, rank, url, title, date_time) in self.result:
            time_str = date_time.strftime('%Y/%m/%d')
            ws.append((domain, keyword, self.ruler.engine_name, page, rank, url, title, time_str))
        wb.save(file_name)
        self.result = []
        print('查询结束，查询结果保存在 %s' % file_name)
        self.save_un_searched()
        self.save_error_list(self.error_list)
        self.save_unsafe_item_list()

    def save_un_searched(self):
        un_searched_keywords = []
        for keyword in self.keyword_domains_map.keys():
            if keyword not in self.searched_keywords:
                un_searched_keywords.append(keyword)
        if len(un_searched_keywords) != 0:
            file_name = '未查找关键词-%s.xlsx' % get_cur_time_filename()
            wb = Workbook()
            ws = wb.active
            for keyword in un_searched_keywords:
                ws.append((keyword,))
            wb.save(file_name)
            print('未查询结果保存在 %s' % file_name)

    def save_error_list(self, err_list):
        if len(err_list) == 0:
            return
        filename = f'排名查询过程中产生的错误-{self.ruler.engine_name}-{get_cur_time_filename()}.log'
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('\n\n'.join(err_list))
        print(filename)
        print('排名查询过程中产生了一些错误，虽然没有终止运行，但是可能会让结果不够准确，请将记录发给开发人员')

    def save_unsafe_item_list(self):
        if len(self.unsafe_item_list) == 0:
            return
        filename = f'关键词是否空白以及安全提醒网站-{self.ruler.engine_name}-{get_cur_time_filename()}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.append(('关键词', '是否空白', '安全提醒', '页数', '排名'))
        for (keyword, has_no_result, unsafe_url, page, rank) in self.unsafe_item_list:
            ws.append((keyword, has_no_result, unsafe_url, page, rank))
        wb.save(filename)


class SiteSpider(Spider):
    def __init__(self, ruler_class):
        Spider.__init__(self, ruler_class)
        self.domain_titles_map = {}
        self.main()

    def search(self):
        self.started = True
        start_time = datetime.now()
        self.domain_titles_map = {}
        domain_set = self.get_input()
        for domain in domain_set:
            self.get_domain(domain)
        self.save_result()
        end_time = datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))
        self.started = False

    def get_input(self):
        path = '.\\要查收录的网址列表XLSX'
        domain_set = set()
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            wb = load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                domain_set.add(row[0])
        return domain_set

    def get_domain(self, domain):
        print('开始查找的域名为 %s' % domain)
        self.domain_titles_map[domain] = []
        page = 1
        soup = self.get_page(domain, page, None)
        page_url = self.ruler.get_next_page_url(soup)
        while soup and self.ruler.has_next_page(soup):
            page += 1
            soup = self.get_page(domain, page, page_url)
            page_url = self.ruler.get_next_page_url(soup)

    def get_page(self, domain, page, page_url):
        print('开始第%d页' % page)
        params = self.ruler.get_params('site:%s' % domain, page)
        if page_url:
            (r, soup, all_item) = self.safe_request(page_url)
        else:
            (r, soup, all_item) = self.safe_request(self.ruler.base_url, params=params)
        for item in all_item:
            self.domain_titles_map[domain].append(self.ruler.get_title(item))
        return soup

    def save_result(self):
        if not self.started:
            return
        wb = Workbook()
        for domain in self.domain_titles_map.keys():
            ws = wb.create_sheet(domain, 0)
            for title in self.domain_titles_map[domain]:
                ws.append((title,))
        self.domain_titles_map = {}
        file_name = '收录标题-%s-%s.xlsx' % (self.ruler.engine_name, get_cur_time_filename())
        wb.save(file_name)


class CheckSpider(Spider):
    def __init__(self, ruler_class):
        Spider.__init__(self, ruler_class)
        self.main()

    def search(self):
        self.started = True
        start_time = datetime.now()
        prices = self.get_input()
        keyword_domains_map = self.get_keyword_domain_map(prices)
        ranks = self.get_ranks(self.ruler, keyword_domains_map)
        self.check_price(prices, ranks)
        end_time = datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))
        self.started = False

    def save_result(self):
        pass

    def get_input(self):
        file_path = ''
        import_dir = '报价'
        path = '.\\%s' % import_dir
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            break
        if file_path == '' or not file_path.endswith('.xlsx'):
            raise MyError('%s目录之下没有发现xlsx文件' % import_dir)
        wb = load_workbook(file_path)
        return wb.active

    def get_keyword_domain_map(self, prices):
        keyword_domains_map = {}
        if self.is_keyword_domain_map:
            keywords = set()
            domains = set()
            for (index, keyword, domain, exponent, price3, price5, rank, charge) \
                    in prices.iter_rows(min_row=2, values_only=True):
                if index is not None:
                    keywords.add(keyword)
                    domains.add(domain)
            for keyword in keywords:
                keyword_domains_map[keyword] = domains
        else:
            for (index, keyword, domain, exponent, price3, price5, rank, charge) \
                    in prices.iter_rows(min_row=2, values_only=True):
                if index is not None:
                    if keyword not in keyword_domains_map.keys():
                        keyword_domains_map[keyword] = set()
                    keyword_domains_map[keyword].add(domain)
        return keyword_domains_map

    def get_ranks(self, ruler, keyword_domains_map):
        results, error_list = LittleRankSpider(self).get_ranks(ruler, keyword_domains_map, 1)
        self.save_error_list(error_list)
        ranks = {}
        for (domain, keyword, page, rank, _, _, _) in results:
            if keyword not in ranks.keys():
                ranks[keyword] = {}
            rank = (page - 1) * 10 + rank
            if domain not in ranks[keyword].keys() \
                    or rank < ranks[keyword][domain]:
                ranks[keyword][domain] = rank
        return ranks

    def check_price(self, prices, ranks):
        total_price = 0
        wb = Workbook()
        ws = wb.active
        ws.append(('序号', '关键词', '网址', '指数', '前三名价格', '四、五名价格', '当前排名', '今日收费', '核对排名', '核对收费'))
        for (index, keyword, domain, exponent, price3, price5, rank, charge) \
                in prices.iter_rows(min_row=2, values_only=True):
            if index is not None:
                check_rank = self.get_rank(ranks, keyword, domain)
                check_price = self.get_price(check_rank, price3, price5)
                total_price = total_price + check_price
                ws.append((index, keyword, domain, exponent, price3, price5, rank, charge, check_rank, check_price))
        ws.append((None, None, None, None, None, None, None, None, '核对总价', total_price))
        file_name = '核对结果-%s-%s.xlsx' % (self.ruler.engine_name, get_cur_time_filename())
        wb.save(file_name)
        input('核对完毕，核对结果保存在%s' % file_name)

    def get_rank(self, ranks, keyword, domain):
        return ranks.get(keyword, {}).get(domain, 0)

    def get_price(self, rank, price3, price5):
        if rank <= 0 or rank > 5:
            return 0
        if rank <= 3:
            return float(price3)
        if rank <= 5:
            return float(price5)

    def save_error_list(self, err_list):
        if len(err_list) == 0:
            return
        filename = f'核对过程中产生的错误-{self.ruler.engine_name}-{get_cur_time_filename()}.log'
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('\n\n'.join(err_list))
        print(filename)
        print('核对过程中产生了一些错误，虽然没有终止运行，但是可能会让结果不够准确，请将记录发给开发人员')
