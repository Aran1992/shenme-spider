import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import os
import time
import datetime
import traceback
from urllib.parse import urlparse
from configparser import ConfigParser


class MyError(RuntimeError):
    pass


def get_cur_time_filename():
    return time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())


def format_cd_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d小时%02d分%02d秒" % (h, m, s)


PAGE = 10
TIMEOUT = 1

global_url = ''
global_text = ''
global_file_name = ''


def search():
    start_time = datetime.datetime.now()
    (keyword_set, domain_set) = get_input()
    init_output()
    print('总共要查找%s关键词，有%s个网站' % (len(keyword_set), len(domain_set)))
    for i, keyword in enumerate(keyword_set):
        get_rank(i + 1, keyword, domain_set)
    print('查询结束，查询结果保存在 %s' % global_file_name)
    end_time = datetime.datetime.now()
    print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))


def start():
    now = datetime.datetime.now()
    cfg = ConfigParser()
    cfg.read('config.ini')
    hour = int(cfg.get('config', 'hour'))
    x = datetime.datetime(now.year, now.month, now.day, hour)
    if x <= now:
        x = datetime.datetime.fromtimestamp(x.timestamp() + 24 * 60 * 60)
    wait_time = (x - now).total_seconds()
    print('下次查询时间为%s，将在%s后开始' % (x, format_cd_time(wait_time)))
    time.sleep(wait_time)
    search()
    start()


def init_output():
    global global_file_name
    global_file_name = '关键词排名-%s.xlsx' % get_cur_time_filename()
    wb = Workbook()
    ws = wb.active
    ws.append(('域名', '关键词', '搜索引擎', '排名', '真实地址', '标题', '查询时间'))
    wb.save(global_file_name)


def get_input():
    # todo 没有对应文件的时候进行提示
    file_path = ''
    path = '.\\import'
    for file in os.listdir(path):
        file_path = os.path.join(path, file)
        break
    if file_path == '' or not file_path.endswith('.xlsx'):
        raise MyError('import目录之下没有发现xlsx文件')
    wb = load_workbook(file_path)
    ws = wb.active
    k = set()
    d = set()
    # todo 文档没有按照格式进行填写的时候进行提示
    for (domain, keyword) in ws.iter_rows(min_row=2, values_only=True):
        k.add(keyword)
        d.add(domain)
    return k, d


def get_rank(index, keyword, domain_set):
    print('开始抓取第%s个关键词：%s' % (index, keyword))
    for i in range(PAGE):
        get_page(i + 1, keyword, domain_set)


def get_page(page, keyword, domain_set):
    global global_url, global_text
    print('开始第%d页' % page)
    # todo 这个网址这样请求就一定能够返回想要的结果吗？
    # https://m.sm.cn/s?q=%E7%99%BE%E5%BA%A6&page=2&by=next&from=smor&tomode=center&safe=1
    params = {
        'q': keyword,
        'page': page,
        'by': 'next',
        'from': 'smor',
        'tomode': 'center',
        'safe': '1',
    }
    # todo 不同的UA有什么影响
    # todo 要怎么模拟真实的用户进行搜索 HEAD要怎么填写
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1'
        # 'User-Agent': 'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Mobile Safari/537.36'
    }
    # todo 网络出现问题的时候怎么办？
    # todo 抓取的内容有问题的时候怎么办？
    r = None
    soup = None
    while r is None or soup is None:
        try:
            r = requests.get('https://m.sm.cn/s', params=params, headers=headers)
        except (requests.exceptions.ConnectionError, requests.exceptions.ChunkedEncodingError):
            print('检查到网络断开，%s秒之后尝试重新抓取' % TIMEOUT)
            time.sleep(TIMEOUT)
            continue
        global_url = r.url
        global_text = r.text
        soup = BeautifulSoup(r.text, 'lxml')
        if soup.body is None:
            print('请求到的页面的内容为空，将再次进行请求')
            soup = None
    all_item = get_all_item(soup)
    result = []
    rank = 1
    for item in all_item:
        link = item.find('a')
        url = link.get('href')
        d = get_url_domain(url)
        if d in domain_set:
            result.append((
                d,
                keyword,
                '%d-%d' % (page, rank),
                url,
                get_title(item),
                datetime.datetime.now()
            ))
        rank += 1
    set_output(result)


def get_url_domain(url):
    li = urlparse(url).netloc.split('.')
    length = len(li)
    return '{}.{}'.format(li[length - 2], li[length - 1])


def get_all_item(soup):
    return soup.find_all('div', class_='ali_row')


def get_title(item):
    return ''.join(item.find('a').findAll(text=True))


def set_output(result):
    global global_file_name
    wb = load_workbook(filename=global_file_name)
    ws = wb.active
    for (domain, keyword, rank, url, title, date_time) in result:
        ws.append((domain, keyword, '神马', rank, url, title, date_time))
    wb.save(global_file_name)


def main():
    try:
        while True:
            mode = input('定时运行（输入1）还是马上运行（输入0）？')
            if mode == '0':
                search()
            elif mode == '1':
                start()
            else:
                print('输入了未知模式，请重新输入')
    except MyError as e:
        input(e)
    except KeyboardInterrupt:
        input('已经强行退出程序')
    except:
        filename = 'error-%s.log' % get_cur_time_filename()
        f = open(filename, 'w', encoding='utf-8')
        f.write('''%s

    请求的URL为：
    %s

    返回的内容为：
    %s
    ''' % (traceback.format_exc(), global_url, global_text))
        f.close()
        traceback.print_exc()
        input('请将最新的error.log文件发给技术人员')


if __name__ == '__main__':
    main()
