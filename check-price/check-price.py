from openpyxl import load_workbook, Workbook
import os
import time
import traceback


def get_cur_time_filename():
    return time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())


class MyError(RuntimeError):
    pass


class Check:
    def __init__(self):
        self.rank_dict = {}
        result = self.load_workbook('查询结果')
        for (domain, keyword, rank) in result.iter_rows(min_row=2, values_only=True):
            if keyword not in self.rank_dict.keys():
                self.rank_dict[keyword] = {}
            self.rank_dict[keyword][domain] = rank

        price = self.load_workbook('报价')
        total_price = 0

        wb = Workbook()
        ws = wb.active
        ws.append(('序号', '关键词', '网址', '指数', '前三名价格', '四、五名价格', '当前排名', '今日收费', '核对排名', '核对收费'))
        for (index, keyword, domain, exponent, price3, price5, rank, charge) \
                in price.iter_rows(min_row=2, values_only=True):
            check_rank = self.get_rank(keyword, domain)
            check_price = self.get_price(check_rank, price3, price5)
            total_price = total_price + check_price
            ws.append((index, keyword, domain, exponent, price3, price5, rank, charge, check_rank, check_price))
        ws.append((None, None, None, None, None, None, None, None, None, total_price))
        file_name = '核对结果-%s.xlsx' % get_cur_time_filename()
        wb.save(file_name)
        input('核对完毕，核对结果保存在%s' % file_name)

    def load_workbook(self, import_dir):
        file_path = ''
        path = '.\\%s' % import_dir
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            break
        if file_path == '' or not file_path.endswith('.xlsx'):
            raise MyError('%s目录之下没有发现xlsx文件' % import_dir)
        wb = load_workbook(file_path)
        return wb.active

    def get_rank(self, keyword, domain):
        return self.rank_dict.get(keyword, {}).get(domain, 0)

    def get_price(self, rank, price3, price5):
        if rank <= 0 or rank > 5:
            return 0
        if rank <= 3:
            return price3
        if rank <= 5:
            return price5


if __name__ == '__main__':
    try:
        Check()
    except:
        with open('error-%s.log' % get_cur_time_filename(), 'w', encoding='utf-8') as f:
            f.write(traceback.format_exc())
            input('请将最新的error.log文件发给技术人员')
