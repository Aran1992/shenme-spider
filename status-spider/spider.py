import asyncio
import datetime
import os
import time
import traceback
from configparser import ConfigParser

import aiohttp
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# import this seems unused but it's to prevent 'LookupError: unknown encoding: idna'
import encodings.idna

HEADERS = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;'
              'q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Host': 'www.rfid808.com',
    'Pragma': 'no-cache',
    'Sec-Fetch-Site': 'none',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Mobile Safari/537.36',
}


def create_headers(site):
    HEADERS['Host'] = adjust_site(site)
    return HEADERS


def get_cur_time_filename():
    return time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())


def format_cd_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d小时%02d分%02d秒" % (h, m, s)


def format_error(tag, e):
    return f'{tag} ({type(e).__name__}: {str(e)})'


def adjust_site(url):
    if url.startswith('www.'):
        return url
    else:
        return f'www.{url}'


class MyError(RuntimeError):
    pass


class StatusSpider:
    def __init__(self):
        self.results = {}
        self.url_list = []
        self.wait_url_list = []
        self.complete_count = 0
        self.cfg = ConfigParser()
        self.cfg.read('config.ini')
        self.max_count = int(self.cfg.get('config', 'max_count'))
        self.timeout = float(self.cfg.get('config', 'timeout'))
        self.search_included = self.cfg.get('config', 'search_included') == '1'
        self.search_http = self.cfg.get('config', 'search_http') == '1'
        self.search_https = self.cfg.get('config', 'search_https') == '1'
        self.search_keywords = self.cfg.get('config', 'search_keywords') == '1'
        self.search_generator = self.cfg.get('config', 'search_generator') == '1'
        self.search_refresh_datetime = self.cfg.get('config', 'search_refresh_datetime') == '1'
        self.main()
        input()

    def main(self):
        try:
            self.get_mode()
        except MyError as e:
            self.save_result()
            input(e)
        except KeyboardInterrupt:
            self.save_result()
            input('已经强行退出程序')
        except:
            self.save_result()
            filename = 'error-%s.log' % get_cur_time_filename()
            with open(filename, 'w', ) as f:
                f.write(traceback.format_exc())
            traceback.print_exc()
            input('请将最新的error.log文件发给技术人员')

    def get_mode(self):
        run_mode = self.cfg.get('config', 'schedule')
        if run_mode == '0':
            asyncio.run(self.search())
        elif run_mode == '1':
            self.start()
        else:
            print('输入了未知模式，请重新输入')

    def start(self):
        now = datetime.datetime.now()
        hour = int(self.cfg.get('config', 'hour'))
        start_time = datetime.datetime(now.year, now.month, now.day, hour)
        if start_time <= now:
            start_time = datetime.datetime.fromtimestamp(start_time.timestamp() + 24 * 60 * 60)
        wait_time = (start_time - now).total_seconds()
        print('下次查询时间为%s，将在%s后开始' % (start_time, format_cd_time(wait_time)))
        time.sleep(wait_time)
        asyncio.run(self.search())
        self.start()

    async def search(self):
        start_time = datetime.datetime.now()
        self.results = {}
        self.url_list = self.get_input()
        self.wait_url_list = self.url_list[self.max_count:]
        self.complete_count = 0
        async with aiohttp.ClientSession() as session:
            await asyncio.gather(*[self.get_url_status(session, url) for url in self.url_list[:self.max_count]])
        self.save_result()
        end_time = datetime.datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))

    def get_input(self):
        url_list = []
        path = '.\\import'
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            wb = load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[0]:
                    url_list.append(row[0])
        return url_list

    async def get_url_status(self, session, url):
        print(f'开始查询 {url} 状态')
        if url not in self.results:
            self.results[url] = {}
        tasks = []
        # 如果既不查询http状态也不查询https状态 那么就要通过http请求获得页面 来获得剩余的信息
        if self.search_http or (not self.search_https
                                and self.search_keywords or self.search_generator or self.search_refresh_datetime):
            tasks.append(asyncio.create_task(self.get_url(session, url, 'http')))
        if self.search_https:
            tasks.append(asyncio.create_task(self.get_url(session, url, 'https')))
        if self.search_included:
            tasks.append(asyncio.create_task(self.is_site_included(session, url)))
        await asyncio.gather(*tasks)
        self.complete_count = self.complete_count + 1
        print(f'{url} 状态查询结束 还剩{len(self.url_list) - self.complete_count}个正在查询')
        if len(self.wait_url_list) != 0:
            url = self.wait_url_list.pop(0)
            await self.get_url_status(session, url)

    async def get_url(self, session, url, protocol):
        result = self.results[url]
        try:
            async with session.get(f'{protocol}://{adjust_site(url)}',
                                   headers=create_headers(url),
                                   timeout=aiohttp.ClientTimeout(total=self.timeout)) as resp:
                result[protocol] = resp.status
                if not self.is_all_info_collected(result):
                    soup = BeautifulSoup(await resp.text(), 'lxml')
                    keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
                    if keywords_meta:
                        result['keywords'] = keywords_meta.attrs['content']
                    else:
                        result['keywords'] = '没有包含keyword的<meta>标签'
                    generator_meta = soup.find('meta', attrs={'name': 'generator'})
                    result['generator'] = generator_meta and 'wp' or 'zm'
                    suffix = result['generator'] == 'wp' and 'feed' or 'rss.php'
                    if self.search_refresh_datetime:
                        try:
                            async with session.get(f'{protocol}://{adjust_site(url)}/{suffix}',
                                                   headers=create_headers(url),
                                                   timeout=aiohttp.ClientTimeout(total=self.timeout)) as resp2:
                                soup2 = BeautifulSoup(await resp2.text(), 'lxml')
                                pub_date = (soup2.find('pubDate') or soup2.find('pubdate'))
                                if pub_date:
                                    dt_str = pub_date.get_text()
                                    if '+' in dt_str:
                                        dt_str = dt_str.split('+')[0].strip()
                                        dt = datetime.datetime.strptime(dt_str, '%a, %d %b %Y %H:%M:%S')
                                    else:
                                        dt = datetime.datetime.strptime(dt_str.strip(), '%Y-%m-%d %H:%M:%S')
                                    result['refresh_datetime'] = dt
                                else:
                                    result['refresh_datetime'] = '未找到<pubdate>元素'
                        except asyncio.TimeoutError:
                            result['refresh_datetime'] = '查询超时'
                        except Exception as e:
                            result['refresh_datetime'] = format_error('查询出错', e)
        except asyncio.TimeoutError:
            result[protocol] = '请求超时'
        except UnicodeDecodeError as e:
            result[protocol] = format_error('网站编码过于奇特，没办法解析', e)
        except aiohttp.client_exceptions.ClientConnectorCertificateError as e:
            result[protocol] = format_error('SSL证书验证失败', e)
        except (
                aiohttp.client_exceptions.ClientConnectorError,
                aiohttp.client_exceptions.ClientOSError,
                aiohttp.client_exceptions.ServerDisconnectedError,
        ) as e:
            result[protocol] = format_error('查询出错', e)

    async def is_site_included(self, session, url):
        params = {'word': f'site:${url}'}
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/75.0.3770.100 '
                          'Safari/537.36'
        }
        try:
            async with session.get(
                    'https://www.baidu.com/s',
                    params=params,
                    headers=headers,
                    timeout=aiohttp.ClientTimeout(total=self.timeout)
            ) as resp:
                soup = BeautifulSoup(await resp.text(), 'lxml')
                div_root = soup.find('div', id='content_left')
                if div_root:
                    includes = div_root.find_all('div', recursive=False, id=lambda id_: id_ != 'rs_top_new')
                    self.results[url]['included'] = len(includes) != 0
                else:
                    self.results[url]['included'] = False
        except asyncio.TimeoutError:
            self.results[url]['included'] = '请求超时'

    def save_result(self):
        print('开始保存查询结果')
        name_flag_key_tuple_list = [
            ('百度是否收录', self.search_included, 'included',),
            ('HTTP', self.search_http, 'http',),
            ('HTTPS', self.search_https, 'https',),
            ('关键词', self.search_keywords, 'keywords',),
            ('模板', self.search_generator, 'generator',),
            ('更新时间', self.search_refresh_datetime, 'refresh_datetime',),
        ]

        wb = Workbook()
        ws = wb.active

        row = ['网站', ]
        for (name, search, _) in name_flag_key_tuple_list:
            if search:
                row.append(name)
        ws.append(tuple(row))

        for url in self.url_list:
            if url in self.results:
                item = self.results[url]

                if 'included' in item:
                    if item['included']:
                        item['included'] = '是'
                    else:
                        item['included'] = '否'

                if (self.search_http and item['http'] != 200) and (self.search_https and item['https'] != 200):
                    for key in ['keywords', 'generator', 'refresh_datetime', ]:
                        if key not in item:
                            item[key] = '由于无法请求到页面，所以未能查询到信息'

                row = [url]
                for (_, search, key) in name_flag_key_tuple_list:
                    if search:
                        row.append((key in item) and item[key] or '由于未知原因，未能查询到信息')
                ws.append(tuple(row))
            else:
                row = [url]
                for (_, search, _) in name_flag_key_tuple_list:
                    if search:
                        row.append('由于强行终止程序或者发生异常，未能进行查询')
                ws.append(tuple(row))
        file_name = f'状态查询-{get_cur_time_filename()}.xlsx'
        wb.save(file_name)
        print(f'查询结果保存在 {file_name}')

    def is_all_info_collected(self, result):
        return (not self.search_http or ('http' in result)) \
               and (not self.search_https or ('https' in result)) \
               and (not self.search_keywords or ('keywords' in result)) \
               and (not self.search_generator or ('generator' in result)) \
               and (not self.search_refresh_datetime or ('refresh_datetime' in result))

    def get_remain_count(self):
        total = len(self.url_list)
        complete = 0
        for key in self.results.keys():
            if self.is_all_info_collected(self.results[key]):
                complete = complete + 1
        return total - complete


if __name__ == '__main__':
    StatusSpider()
