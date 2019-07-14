import requests
from bs4 import BeautifulSoup, Comment
from openpyxl import load_workbook, Workbook
import os
import time
import datetime
import traceback
from urllib.parse import urlparse, urljoin
from configparser import ConfigParser

PAGE = 10
TIMEOUT = 1


class MyError(RuntimeError):
    pass


def get_cur_time_filename():
    return time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime())


def format_cd_time(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d小时%02d分%02d秒" % (h, m, s)


class Spider:
    def __init__(self):
        self.file_name = ''
        self.url = ''
        self.text = ''
        try:
            while True:
                mode = input('定时运行（输入1）还是马上运行（输入0）？')
                if mode == '0':
                    self.search()
                elif mode == '1':
                    self.start()
                else:
                    print('输入了未知模式，请重新输入')
        except MyError as e:
            input(e)
        except KeyboardInterrupt:
            input('已经强行退出程序')
        except:
            filename = 'error-%s.log' % get_cur_time_filename()
            f = open(filename, 'w', encoding='utf-8')
            f.write('''%s

请求的URL为：
%s

返回的内容为：
%s
''' % (traceback.format_exc(), self.url, self.text))
            f.close()
            traceback.print_exc()
            input('请将最新的error.log文件发给技术人员')

    def start(self):
        now = datetime.datetime.now()
        cfg = ConfigParser()
        cfg.read('config.ini')
        hour = int(cfg.get('config', 'hour'))
        x = datetime.datetime(now.year, now.month, now.day, hour)
        if x <= now:
            x = datetime.datetime.fromtimestamp(x.timestamp() + 24 * 60 * 60)
        wait_time = (x - now).total_seconds()
        print('下次查询时间为%s，将在%s后开始' % (x, format_cd_time(wait_time)))
        time.sleep(wait_time)
        self.search()
        self.start()

    def search(self):
        start_time = datetime.datetime.now()
        self.init_output()
        (keyword_set, domain_set) = self.get_input()
        print('总共要查找%s关键词，有%s个网站' % (len(keyword_set), len(domain_set)))
        for i, keyword in enumerate(keyword_set):
            self.get_rank(i + 1, keyword, domain_set)
        print('查询结束，查询结果保存在 %s' % self.file_name)
        end_time = datetime.datetime.now()
        print('本次查询用时%s' % format_cd_time((end_time - start_time).total_seconds()))

    def init_output(self):
        self.file_name = '关键词排名-%s.xlsx' % get_cur_time_filename()
        wb = Workbook()
        ws = wb.active
        ws.append(('域名', '关键词', '搜索引擎', '排名', '真实地址', '标题', '查询时间'))
        wb.save(self.file_name)

    def get_input(self):
        file_path = ''
        path = '.\\import'
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            break
        if file_path == '' or not file_path.endswith('.xlsx'):
            raise MyError('import目录之下没有发现xlsx文件')
        wb = load_workbook(file_path)
        ws = wb.active
        k = set()
        d = set()
        for (domain, keyword) in ws.iter_rows(min_row=2, values_only=True):
            k.add(keyword)
            d.add(domain)
        return k, d

    def get_rank(self, index, keyword, domain_set):
        print('开始抓取第%s个关键词：%s' % (index, keyword))
        for i in range(PAGE):
            self.get_page(i + 1, keyword, domain_set)

    def get_page(self, page, keyword, domain_set):
        print('开始第%d页' % page)
        # https://wap.sogou.com/web/search/ajax_query.jsp?type=1&uID=AAEoNfGfKAAAAAqPLE5KcQAA1wA=&v=5&dp=1&pid=sogou-waps-7880d7226e872b77&rcer=hNz_aRIBWIwCGa7H&keyword=%E6%97%85%E6%B3%95%E5%B8%88%E8%90%A5%E5%9C%B0&suuid=f62fab99-2cd6-4d62-a9d1-8b400c6729d4&p=2&s_from=pagenext&showextquery=1&IPLOC=&insite=
        params = {
            'keyword': keyword,
            'p': page,
        }
        headers = {
            'User-Agent': 'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Mobile Safari/537.36'
        }
        (r, soup) = self.save_request('http://wap.sogou.com/web/search/ajax_query.jsp', params=params, headers=headers)
        all_item = self.get_all_item(soup)
        result = []
        rank = 1
        for item in all_item:
            url = item.get('href')
            if url.startswith('javascript'):
                continue
            if not url.startswith('http'):
                url = urljoin(r.url, url)
                r = requests.get(url)
                if r.url.startswith('http://wap.sogou.com/web/search/'):
                    sub_soup = BeautifulSoup(r.text, 'lxml')
                    btn = sub_soup.find('div', class_='btn')
                    link = btn.find('a')
                    url = link.get('href')
                else:
                    url = r.url
            d = self.get_url_domain(url)
            if d in domain_set:
                result.append((
                    d,
                    keyword,
                    '%d-%d' % (page, rank),
                    url,
                    self.get_title(item),
                    datetime.datetime.now()
                ))
            rank += 1
        self.set_output(result)

    def get_all_item(self, soup):
        return soup.find_all('a', class_='resultLink')

    def get_url_domain(self, url):
        li = urlparse(url).netloc.split('.')
        length = len(li)
        return '{}.{}'.format(li[length - 2], li[length - 1])

    def get_title(self, item):
        return ''.join(item.findAll(text=lambda text: not isinstance(text, Comment)))

    def set_output(self, result):
        wb = load_workbook(filename=self.file_name)
        ws = wb.active
        for (domain, keyword, rank, url, title, date_time) in result:
            ws.append((domain, keyword, '搜狗MOBILE', rank, url, title, date_time))
        wb.save(self.file_name)

    def save_request(self, url, *, params, headers):
        r = None
        soup = None
        while r is None or soup is None:
            try:
                r = requests.get(url, params=params, headers=headers)
            except (requests.exceptions.ConnectionError, requests.exceptions.ChunkedEncodingError):
                print('检查到网络断开，%s秒之后尝试重新抓取' % TIMEOUT)
                time.sleep(TIMEOUT)
                continue
            self.url = r.url
            self.text = r.text
            soup = BeautifulSoup(r.text, 'lxml')
            if soup.body is None:
                print('请求到的页面的内容为空，将再次进行请求')
                soup = None
        return r, soup


if __name__ == '__main__':
    Spider()
